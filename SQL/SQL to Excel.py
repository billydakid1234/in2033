# coding=utf-8
import sqlite3
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

# -----------------------------
# SETTINGS
# -----------------------------
sqlite_db_path = "For Demo SQL/CA_db.db"
excel_output_path = "CA_db_demo.xlsx"

# -----------------------------
# CONNECT TO SQLITE
# -----------------------------
conn = sqlite3.connect(sqlite_db_path)
cursor = conn.cursor()

# Get all user tables
cursor.execute("""
    SELECT name
    FROM sqlite_master
    WHERE type='table'
      AND name NOT LIKE 'sqlite_%'
    ORDER BY name;
""")
tables = [row[0] for row in cursor.fetchall()]

if not tables:
    print("No tables found in the SQLite database.")
    conn.close()
    exit()

# -----------------------------
# CREATE EXCEL WORKBOOK
# -----------------------------
wb = Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

for table_name in tables:
    print(f"Exporting table: {table_name}")

    # Read full table
    cursor.execute(f'SELECT * FROM "{table_name}"')
    rows = cursor.fetchall()

    # Get column names
    column_names = [description[0] for description in cursor.description]

    # Create worksheet
    ws = wb.create_sheet(title=table_name[:31])  # Excel sheet names max 31 chars

    # Write headers
    ws.append(column_names)

    # Write data rows
    for row in rows:
        ws.append(row)

    # Create Excel table if there is at least a header row
    if len(column_names) > 0:
        last_row = ws.max_row
        last_col = ws.max_column
        table_range = f"A1:{get_column_letter(last_col)}{last_row}"

        excel_table = Table(
            displayName=f"Table_{table_name[:20].replace(' ', '_')}",
            ref=table_range
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        excel_table.tableStyleInfo = style
        ws.add_table(excel_table)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# Save workbook
wb.save(excel_output_path)

# Close connection
conn.close()

print(f"Done. Excel file saved as: {os.path.abspath(excel_output_path)}")